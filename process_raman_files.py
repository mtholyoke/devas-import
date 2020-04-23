from __future__ import print_function
import numpy as np
from collections import defaultdict
from openpyxl import load_workbook
from os.path import basename
from time import time

from importer import _TrajImporter


class RamanImporter(_TrajImporter):
    driver = None
    file_ext = '.txt'
    pkey_field = 'spectrum_number'

    def parse_masterfile(self, cfile):
        print('Loading masterfile...')
        start = time()
        book = load_workbook(cfile, data_only=True)
        sheet = book.active
        for row in sheet.iter_rows(min_row=1, max_row=1):
            headers = [cell.value for cell in row]
        meta = defaultdict(list)
        id_ind = headers==self.pkey_field
        for i, row in enumerate(sheet.rows):
            if i < 1 or row[id_ind].value is None:
                continue
            for header, cell in zip(headers, row):
                if header is not None:
                    meta[header].append(cell.value)
        print('  done. %.2fs' % (time() - start))
        return meta

    def _process_spectra(self, fname, metadata):
        pkeys = np.array(metadata[self.pkey_field])
        meta_idx, = np.where(pkeys == self._get_id(fname))
        
        if len(meta_idx) != 1:
            print('  Cannot match spectrum and masterfile', fname)
            return
        meta = {key: val[meta_idx[0]] for key, val in metadata.items()}
        try:
            spectrum = np.genfromtxt(fname, delimiter=',')
            if spectrum.ndim != 2 or spectrum.shape[1] != 2:
                raise ValueError('Spectrum must be a trajectory')
        except Exception as e:
            print(' ', e, fname)
            return
        # Make sure wavelengths are increasing
        if spectrum[0,0] > spectrum[1,0]:
            spectrum = spectrum[::-1]
        return spectrum, meta

    def _get_id(self, path):
        try:
            return int(basename(path).rstrip(self.file_ext).split('_')[0])
        except:
            return


if __name__ == '__main__':
    RamanImporter().main()
