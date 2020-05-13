from __future__ import print_function
import numpy as np
import itertools
from collections import defaultdict
from openpyxl import load_workbook
from os.path import basename
from time import time

from importer import _TrajImporter


class MossbauerImporter(_TrajImporter):
    driver = None
    file_ext = '.txt'
    pkey_field = 'Sample #'
    n_chans = 512
    superman_fields = set(['Sample #', 'T(K)', 'Sample Name', 'Post?',
                           'Dana Group', 'Group Folder', 'Owner/Source',
                           # Add 'Pubs' but make a search widget first
                           # Remove 'Group Folder' after Darby edits
                           ])

    def parse_masterfile(self, cfile):
        print('Loading masterfile...')
        start = time()
        book = load_workbook(cfile, data_only=True)
        sheet = book.active
        for row in sheet.iter_rows(min_row=1, max_row=1):
            headers = [cell.value for cell in row]
        meta = defaultdict(list)
        id_ind = headers == self.pkey_field
        for i, row in enumerate(sheet.rows):
            if i < 1 or row[id_ind].value is None:
                continue
            for header, cell in zip(headers, row):
                if header in self.superman_fields:
                    if header == 'Dana Group' and not cell.value:
                        cell.value = 'n/a'
                    meta[header].append(cell.value)
        print('  done. %.2fs' % (time() - start))
        return meta

    def _process_spectra(self, fname, metadata):
        pkeys = np.asarray(metadata[self.pkey_field], dtype=str) #for unicode
        meta_idx, = np.where(pkeys == self._get_id(fname))
        if len(meta_idx) != 1:
            print('  Cannot match spectrum and masterfile', fname)
            return
        meta_idx = meta_idx[0]
        if metadata['Post?'][meta_idx] is None or \
            metadata['Post?'][meta_idx].upper()!='Y':
            return
        meta = {key: val[meta_idx] for key, val in metadata.items()}
        spectrum = []
        with open(fname) as f:
            for line in itertools.islice(f, 10, None):
                line = line.strip()
                try:
                    row = map(float, line.split())
                    if len(row) != 2:
                      print('  Wrong data format in file', fname)
                      return
                    spectrum.append(np.asarray(row, dtype=float))
                except ValueError:
                    pass
        if len(spectrum) != self.n_chans:
            return
        return np.array(spectrum, dtype=float), meta

    def _get_id(self, path):
        return basename(path).split('.')[0]


if __name__ == '__main__':
    MossbauerImporter().main()
