from __future__ import print_function
from collections import defaultdict
from csv import reader
from glob import glob
from openpyxl import load_workbook
import numpy as np
import os


def find_mhc_spectrum_files(input_dir):
    fpattern = os.path.join(input_dir, '*', '*_spect.csv')
    return [f for f in glob(fpattern)
            if '_TI_' not in f.upper() and '_DARK_' not in f.upper()]

def mhc_spectrum_id(fpath):
    name, _ = os.path.splitext(os.path.basename(fpath))
    # Drop the '_spect' suffix.
    name, suffix = name.rsplit('_', 1)
    assert suffix == 'spect', 'Invalid spectrum path: ' + fpath
    return name

def _get_elements(sheet):
    elements = []
    good_cols = []
    for col, _ in enumerate(sheet.columns, start=1):
        check = sheet.cell(row=1, column=col).value
        if not check or 'X' not in check:
            continue
        name = sheet.cell(row=2, column=col).value
        if not name:
            raise ValueError('Bad name for checked column %d' % col)
        elements.append('e_' + name)
        good_cols.append(col)
    return elements, good_cols

def _is_type(x, typ):
    try:
        typ(x)
    except:
        return False
    else:
        return True

def _is_int(x):
    return _is_type(x, int)

def _is_float(x):
    return _is_type(x, float)

def parse_mhc_masterfile(filename):
    book = load_workbook(filename, data_only=True)
    sheet = book.active
    elements, ecols = _get_elements(sheet)
    samples = []
    rock_types = []
    randoms = []
    matrices = []
    dopants = []
    projects = []
    compositions = defaultdict(list)
    for i, _ in enumerate(sheet.rows):
        if i < 4:
            continue
        samples.append(sheet.cell(row=i, column=1).value)
        rock_types.append(sheet.cell(row=i, column=2).value)
        randoms.append(sheet.cell(row=i, column=3).value)
        matrices.append(sheet.cell(row=i, column=4).value)
        dopants.append(sheet.cell(row=i, column=5).value)
        projects.append(sheet.cell(row=i, column=6).value)
        for elem, col in zip(elements, ecols):
            if col < 7:
                continue
            val = sheet.cell(row=i, column=col).value
            if isinstance(val, float) or isinstance(val, long):
                compositions[elem].append(val)
            else:
                compositions[elem].append(np.nan)
    noncomps = [rock_types, randoms, matrices, dopants, projects]
    return samples, compositions, noncomps


META_FIELDS = set(['Carousel', 'Sample', 'Target', 'Location', 'Atmosphere',
                   'LaserAttenuation', 'DistToTarget', 'Date', 'Projects'])

def process_mhc_spectra(fname, n_chans=None):
    with open(fname,'r') as f:
        contents = list(reader(f, quotechar='+'))
    meta = {}
    prepro = True
    # Spectra are assumed to start at first line of comma-separated float/int values
    for i, line in enumerate(contents):
        if line and ':' in line[0]:
            field = line[0].split(':')[0]
            val = line[0].split(':')[1].strip()
            if field == 'Carousels':
                field = 'Carousel'
                val = val.split(' ')[0].strip()
            if field == 'Dates':
                field = 'Date'
                val = val.split(' ')[0].strip()
            if field == 'Locations':
                field = 'Location'
                val = val.split(' ')[0].strip()
            if field == 'Target:' and val.startswith('T'):
                val = val.lstrip('T')
            if field in META_FIELDS:
                meta[field] = val
        elif all(_is_float(item) for item in line):
            break
        elif all(_is_int(item) for item in line):
            prepro = False  # it's formatted
            break
    if meta['Sample'].lower() in ('ti', 'dark'):
        return
    try:
        data = np.array(contents[i:], dtype=float)
    except:
        print('WARNING: Bad spectra in file:', fname)
        return
    if n_chans and data.shape[0] != n_chans:
        print('WARNING: Wrong number of channels in file:', fname)
        print('Expected', n_chans, 'but got', data.shape[0])
        return
    return data.T, meta, prepro
