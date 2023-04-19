#!/usr/bin/env python3

import csv
import numpy as np
import openpyxl
import os
from collections import defaultdict
from glob import glob
from openpyxl import load_workbook
from time import time


def can_be_float(string):
    """
    Determines if a given string can become a float. 
   
    Parameters
    ----------
    string
        String to be tested.
    
    Returns
    -------
        Boolean reflecting whether the string can be a float.
    """
    try:
        float(string)
    except Exception:
        return False
    else:
        return True


def clean_data(string, cast=int, default=0):
    """
    Takes a string and strips it of non-numerical characters. 

    Parameters
    ----------
    string 
        A metadata value that should be numeric.
    cast 
        What numeric type (int or float) to return.
    default
        Default value if string is empty after cleanup.

    Returns
    -------
        Cleaned string cast as given datatype. 
    """
    charset = [str(x) for x in range(0, 10)] + ['-']
    if cast is float:
        charset.append('.')
    for char in string:
        if char not in charset:
            string = string.replace(char, '')
    if string == '':
        string = default
    return cast(string)


def find_spectrum_files(input_dir, file_ext):
    """
    Scans the data directory for particular file type and return a list of those files.
    Used primarily to scan for *_spect.csv. Files must be in 1 level of 
    subdirectory.

    Parameters
    ---------- 
    input_dir 
        The directory to scan through.
    file_ext 
        The file extension to be searched for.

    Returns
    -------
        A list of files to process. 
    """
    fpattern = os.path.join(input_dir, '*', f'*{file_ext}')
    return [f for f in glob(fpattern)
            if '_TI_' not in f.upper() and '_DARK_' not in f.upper()]


def get_directory(filepath):
    """
    Finds the directory a given file is in. 

    Parameters
    ----------
    filepath
        The full path of a file. 

    Returns
    ------- 
        The name of the directory a file is in.
    """
    return os.path.basename(os.path.dirname(filepath))


def get_element_columns(sheet):
    """
    Gets the element columns from the metadata sheet (LIBS).
    Parameters
    ----------
    sheet
        the metadata sheet.

    Returns
    -------
    elem_cols
        an array of elements from the sheet.
    
    """
    elem_cols = []
    for col, _ in enumerate(sheet.columns, start=1):
        check = sheet.cell(row=1, column=col).value
        if not check or 'X' not in check:
            continue
        name = sheet.cell(row=2, column=col).value
        if not name:
            raise ValueError('Bad name for checked column %d' % col)
        elem_cols.append(('e_' + name, col))
    return elem_cols


def get_spectrum_id(filename):
    """
    Truncates "_spect.csv" from the filename to get ID. Note:
    is explicitly for that, not other file types.

    Parameters
    ----------
    filename
        a string representing the full name of the file.

    Returns
    -------
    name[:-6]
        the name of the file, minus the '_spect.csv' ending.
    
    """
    name, _ = os.path.splitext(os.path.basename(filename))
    name = name.decode() if isinstance(name, bytes) else name
    if not name.endswith('_spect') or len(name) < 7:
        return None
    return name[:-6]


META_FIELDS = ['Carousel', 'Sample', 'Target', 'Location', 'Atmosphere',
               'LaserAttenuation', 'DistToTarget', 'Date', 'Projects']


def load_spectra(filepath, channels=None):
    """
    Loads the spectra from a single spectra file.

    Parameters
    ----------
    filepath
        the full path of a spectra file.
    channels
        number of channels (usually as self.channels).

    Returns
    ------
    data.T
        the spectra of the file. 
    meta
        a dict of meta fields to their values.
    prepro
        a bool is spectra is prepro.
    """
    with open(filepath, 'r') as f:
        contents = list(csv.reader(f, quotechar='+'))
    meta = {}
    prepro = True
    # Spectra are assumed to start at first line of all numeric data
    for i, line in enumerate(contents):
        if isinstance(line, bytes):
            line = line.decode()
        if line and ':' in line[0]:
            field = line[0].split(':')[0]
            val = line[0].split(':')[1].strip()
            if field == 'Carousels':
                field = 'Carousel'
                val = val.split(' ')[0].strip()
            if field == 'Dates':
                field = 'Date'
                val = val.split(' ')[0].strip()
            if field == 'Locations':
                field = 'Location'
                val = val.split(' ')[0].strip()
            if field in META_FIELDS:
                meta[field] = val
        elif all(can_be_float(item) for item in line):
            if '.' not in ''.join(line):
                # All integers means this is formatted.
                prepro = False
            break
    if meta['Sample'].lower() in ('ti', 'dark'):
        return
    # SuperLIBS 10K has a typo in the metadata:
    if meta['Sample'] == 'AGVIA':
        meta['Sample'] = 'AGV1A'
    try:
        data = np.array(contents[i:], dtype=float)
    except Exception as e:
        return f'Bad spectra in file {filepath}: {e}'
    if channels and data.shape[0] != channels:
        e = f'expected {channels} channels, got {data.shape[0]}'
        return f'Wrong channel count in file {filepath}: {e}'
    return data.T, meta, prepro


def parse_millennium_comps(filepath):
    """
    Parses LIBS Millennium_comps file. 

    Parameters
    ----------
    filepath
        the full path to a Millennium-comps file.

    Returns
    -------
    samples
        the names of samples. 
    compositions
        a dictionary of lists with elements as keys.
    noncomps
        all other data within millennium_comps not contained in samples
        or compositions.
    """
    book = openpyxl.load_workbook(filepath, data_only=True)
    sheet = book.active
    elem_cols = get_element_columns(sheet)
    samples = []
    rock_types = []
    randoms = []
    matrices = []
    dopants = []
    projects = []
    compositions = defaultdict(list)
    for i, _ in enumerate(sheet.rows):
        if i < 4:
            continue
        samples.append(sheet.cell(row=i, column=1).value)
        rock_types.append(sheet.cell(row=i, column=2).value)
        randoms.append(sheet.cell(row=i, column=3).value)
        matrices.append(sheet.cell(row=i, column=4).value)
        dopants.append(sheet.cell(row=i, column=5).value)
        projects.append(sheet.cell(row=i, column=6).value)
        for elem, col in elem_cols:
            if col < 7:
                continue
            val = sheet.cell(row=i, column=col).value
            if isinstance(val, float) or isinstance(val, int):
                compositions[elem].append(val)
            else:
                compositions[elem].append(np.nan)
    noncomps = [rock_types, randoms, matrices, dopants, projects]
    return samples, compositions, noncomps

def parse_masterfile(cfile, fields, logger):
    """
    Parses the masterfile for non-libs processors. Includes Raman id 
    modification.

    Parameters
    ----------
    cfile
        the path to the masterfile.
    fields
        superman fields, or pkey depending on Mossbauer vs. Raman.
    logger
        logger passed in from processer for warning purposes.

    Returns
    -------
    meta
        a dict of lists where each meta category maps to all its values
        from the masterfile. 
    """
    # Check if it's raman.
    isMossbauer = fields != 'spectrum_number'
    start = time()
    book = load_workbook(cfile, data_only=True)
    sheet = book.active
    for row in sheet.iter_rows(min_row=1, max_row=1):
        headers = [cell.value for cell in row]
    meta = defaultdict(list)
    if isMossbauer:
        id_ind = headers == next(iter(fields))
    # For Raman, keep fields.
    else:
        id_ind = headers==fields
    for i, row in enumerate(sheet.rows):
        if i < 1 or row[id_ind].value is None:
            continue
        if isMossbauer:
            for header, cell in zip(headers, row):
                if header in fields:
                    if header == 'Dana Group' and not cell.value:
                        cell.value = 'n/a'
                    meta[header].append(cell.value)
        else: 
            val_list = str(row[0].value).split('_')
            val = val_list[0]
            for i in range(1, len(val_list)):
                val = val + "_" + val_list[i]
            row[0].value = val
            for header, cell in zip(headers, row):
                if header == 'spectrum_number':
                    if cell.value in meta[header]:
                        logger.warn('Duplicate value: ', cell.value)
                if header is not None:
                    meta[header].append(cell.value)
    return meta


