import numpy as np
from . import utils
import itertools
from collections import defaultdict
from openpyxl import load_workbook
from os.path import basename
from time import time
from ._base import _TrajectoryProcessor

from _mhc_utils import (
    find_mhc_spectrum_files, mhc_spectrum_id, parse_mhc_masterfile,
    process_mhc_spectra) 


class MossbauerImporter(_TrajectoryProcessor):
    """
    Inherits from importer.py
    Processes spectra data from Mossbauer

    Implements these methods required by importer.py:
    - get_id(filename) returns ID or None
    - process_spectra(filename, metadata) return spectra, meta

    Implements these members required by _base.py:
    - driver: None by default
    - file_ext: '_.txt' by default
    - pkey_field: 'Name' by default

    """
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.logger = self.get_child_logger()
        required = ['channels']
        for attr in required:
            if not hasattr(self, attr):
                raise AttributeError(f'Attribute "{attr}" is required')
        defaults = {
            'driver': 'None',
            'file_ext': '.txt',
            'pkey_field': 'Sample #',
        }
        for key, value in defaults.items():
            if not hasattr(self, key):
                setattr(self, key, value)
        self.superman_fields = set(['Sample #', 'T(K)', 'Sample Name', 'Post?',
                           'Dana Group', 'Group Folder', 'Owner/Source',
                           # Add 'Pubs' but make a search widget first
                           # Remove 'Group Folder' after Darby edits
                           ])

    def get_id(self, filename):
        """
        Returns a string representing the name of an individual spectra file
        without the spectra.csv suffix.
        Parameters filename: the path to a spectra file
        """
        return basename(filename).split('.')[0]
    
    def parse_metadata(self):
        """
        Returns data from Millennium_comps file. 
        """
        self.logger.debug('Parsing metadata')
        return self.parse_masterfile(self.paths['metadata'][0])

    def parse_masterfile(self, cfile):
        print('Loading masterfile...')
        start = time()
        book = load_workbook(cfile, data_only=True)
        sheet = book.active
        for row in sheet.iter_rows(min_row=1, max_row=1):
            headers = [cell.value for cell in row]
        meta = defaultdict(list)
        id_ind = headers == self.pkey_field
        for i, row in enumerate(sheet.rows):
            if i < 1 or row[id_ind].value is None:
                continue
            for header, cell in zip(headers, row):
                if header in self.superman_fields:
                    if header == 'Dana Group' and not cell.value:
                        cell.value = 'n/a'
                    meta[header].append(cell.value)
        print('  done. %.2fs' % (time() - start))
        return meta
        
    def get_input_data(self):
        """
        Returns a dictionary of files in a directory.
        """
        data = {}
        for dd in self.paths['data']:
            files = utils.find_spectrum_files(dd, self.file_ext)
            for file in files:
                if not self.get_id(file):
                    continue
                path = utils.get_directory(file)
                if path not in data:
                    data[path] = []
                data[path].append((self.get_id(file), file))
        return data

    def _write_data(self, fname, meta):
        _TrajectoryProcessor._write_data(self, fname, meta)
        print('  skipped', self.skipped, 'spectra so far because the masterfile disabled them.')

    def process_spectra(self, datafile, metadata):
        """
        Returns a processed single file from a batch of file, 
        including data and metadata

        Parameter datafile: a string representing the path to a spectra file
        """
        pkeys = np.asarray(metadata[self.pkey_field], dtype=str) #for unicode
        meta_idx, = np.where(pkeys == self._get_id(datafile))
        if len(meta_idx) != 1:
            print('  Cannot match spectrum and masterfile', datafile)
            return
        meta_idx = meta_idx[0]
        if metadata['Post?'][meta_idx] is None or \
            metadata['Post?'][meta_idx].upper()!='Y':
            self.skipped += 1
            return
        meta = {key: val[meta_idx] for key, val in metadata.items()}
        spectrum = []
        with open(datafile) as f:
            for line in itertools.islice(f, 10, None):
                line = line.strip()
                try:
                    row = map(float, line.split())
                    if len(row) != 2:
                      print('  Wrong data format in file', datafile)
                      return
                    spectrum.append(np.asarray(row, dtype=float))
                except ValueError:
                    pass
        if len(spectrum) != self.n_chans:
            print('  Expected', self.n_chans, 'channels, got', len(spectrum), 'in', datafile)
            return
        return np.array(spectrum, dtype=float), meta